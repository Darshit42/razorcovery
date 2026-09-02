"""Parse + validate an uploaded contact sheet (CSV or XLSX).

Column detection: we look for canonical fields by fuzzy header match
(see schemas.CANONICAL_FIELDS); an explicit mapping from the UI overrides
it. Phone is required and normalised to +91XXXXXXXXXX. Failure type is
inferred from the sheet or defaulted to payment_retry. Rows that can't be
made into a valid recovery event are returned as errors, not silently
dropped.
"""
from __future__ import annotations

import csv
import io
import re
from typing import Any

from data.schemas import ERROR_CODES
from intake.schemas import CANONICAL_FIELDS, ParsedRow, ParseResult, RowError

_DIGITS = re.compile(r"\D")
_CTRL = re.compile(r"[​-‏‪-‮⁠-⁩﻿]")

FAILURE_TYPES = ("payment_retry", "checkout_abandonment", "mandate_failure")
_FAILURE_ALIASES = {
    "payment": "payment_retry",
    "payment_retry": "payment_retry",
    "retry": "payment_retry",
    "failed_payment": "payment_retry",
    "checkout": "checkout_abandonment",
    "abandonment": "checkout_abandonment",
    "cart": "checkout_abandonment",
    "abandoned": "checkout_abandonment",
    "mandate": "mandate_failure",
    "subscription": "mandate_failure",
    "autopay": "mandate_failure",
    "mandate_failure": "mandate_failure",
}


def normalize_indian_phone(raw: str) -> str | None:
    """-> +91XXXXXXXXXX, or None if it can't be a 10-digit Indian mobile."""
    digits = _DIGITS.sub("", _CTRL.sub("", raw or ""))
    if digits.startswith("91") and len(digits) == 12:
        digits = digits[2:]
    elif digits.startswith("0") and len(digits) == 11:
        digits = digits[1:]
    if len(digits) != 10 or digits[0] not in "6789":
        return None
    return f"+91{digits}"


def _norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (h or "").strip().lower()).strip("_")


def detect_mapping(columns: list[str]) -> dict[str, str | None]:
    norm = {_norm_header(c): c for c in columns}
    out: dict[str, str | None] = {}
    for field, aliases in CANONICAL_FIELDS.items():
        hit = next((norm[a] for a in aliases if a in norm), None)
        if hit is None:  # substring fallback
            hit = next((orig for n, orig in norm.items()
                        if any(a in n for a in aliases)), None)
        out[field] = hit
    return out


def _to_amount(v: Any) -> int | None:
    if v is None:
        return None
    s = re.sub(r"[^\d.]", "", str(v))
    if not s:
        return None
    try:
        n = int(round(float(s)))
        return n if n > 0 else None
    except ValueError:
        return None


def _to_failure_type(v: Any) -> str:
    key = _norm_header(str(v or ""))
    if key in _FAILURE_ALIASES:
        return _FAILURE_ALIASES[key]
    for alias, canon in _FAILURE_ALIASES.items():
        if alias in key:
            return canon
    return "payment_retry"


def _rows_from_csv(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    cols = [c for c in (reader.fieldnames or []) if c is not None]
    rows = [{k: (v or "").strip() for k, v in r.items() if k is not None} for r in reader]
    return cols, rows


def _rows_from_xlsx(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return [], []
    cols = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header)]
    rows = []
    for values in it:
        if values is None or all(v is None for v in values):
            continue
        rows.append({cols[i]: ("" if v is None else str(v).strip())
                     for i, v in enumerate(values) if i < len(cols)})
    wb.close()
    return cols, rows


def parse_sheet(
    data: bytes,
    filename: str,
    *,
    mapping: dict[str, str] | None = None,
    default_failure_type: str | None = None,
) -> ParseResult:
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        columns, rows = _rows_from_xlsx(data)
    else:
        columns, rows = _rows_from_csv(data)

    detected = detect_mapping(columns)
    m = {**detected, **{k: v for k, v in (mapping or {}).items() if v}}

    valid: list[ParsedRow] = []
    errors: list[RowError] = []
    seen: dict[str, int] = {}
    dupes: list[str] = []

    phone_col = m.get("phone")
    if not phone_col:
        errors.append(RowError(0, "phone",
                               "No phone column detected — map one explicitly."))
        return ParseResult(columns, len(rows), [], errors, [],
                           sample=rows[:5], detected_mapping=detected)

    for i, r in enumerate(rows, start=1):
        phone = normalize_indian_phone(r.get(phone_col, ""))
        if not phone:
            errors.append(RowError(i, "phone",
                                   f"Invalid Indian mobile: {r.get(phone_col, '')!r}"))
            continue
        if phone in seen:
            dupes.append(phone)
            errors.append(RowError(i, "phone", f"Duplicate of row {seen[phone]}"))
            continue
        seen[phone] = i

        ftype = (default_failure_type
                 or (_to_failure_type(r.get(m["failure_type"])) if m.get("failure_type")
                     else "payment_retry"))
        if ftype not in FAILURE_TYPES:
            ftype = "payment_retry"

        amount = _to_amount(r.get(m["amount_inr"])) if m.get("amount_inr") else None
        if amount is None:
            errors.append(RowError(i, "amount_inr",
                                   "Missing / unparseable amount (need a positive number)."))
            continue

        name_val = (r.get(m["customer_name"]) or "Customer").strip() if m.get("customer_name") else "Customer"
        ref = (r.get(m["reference_id"]) or "").strip() if m.get("reference_id") else ""
        ref = ref or f"row-{i}"
        err_code = (r.get(m["error_code"]) or "").strip() if m.get("error_code") else ""
        err_code = err_code or ERROR_CODES[ftype][0]
        tz = (r.get(m["timezone"]) or "").strip() if m.get("timezone") else ""
        tz = tz or "Asia/Kolkata"

        valid.append(ParsedRow(
            row_index=i, phone=phone, customer_name=name_val or "Customer",
            amount_inr=amount, failure_type=ftype, reference_id=ref,
            error_code=err_code, timezone=tz, raw=r,
        ))

    return ParseResult(
        columns=columns, total_rows=len(rows), valid_rows=valid, errors=errors,
        duplicate_phones=sorted(set(dupes)), sample=rows[:5], detected_mapping=detected,
    )
